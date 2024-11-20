import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


FILE_PATH = 'Väderdata.xlsx'

def create_excel_file_if_not_exists():
    if not os.path.isfile(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(['Created', 'Longitude', 'Latitude', 'Datum', 'Hour', 'Temperature', 'Rain or snow', 'Provider'])
        
        date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD HH:MM:SS')
        ws.column_dimensions[get_column_letter(1)].width = 20
        wb.add_named_style(date_style)
        wb.save(FILE_PATH)

def append_data_to_excel(data):
    create_excel_file_if_not_exists()
    wb = load_workbook(FILE_PATH)
    ws = wb.active

    for row in data:
        if isinstance(row[0], datetime):
            row[0] = row[0].strftime('%Y-%m-%d %H:%M:%S')

        if isinstance(row[5], float):
            row[5] = int(round(row[5]))

        new_row = ws.append(row)

        temperature = row[5]
        last_row = ws.max_row
        temp_cell = ws.cell(row=last_row, column=6)

        if temperature >= 30:
            temp_cell.fill = PatternFill(start_color="FF5733", end_color="FF5733", fill_type="solid")
        elif 15 <= temperature < 30:
            temp_cell.fill = PatternFill(start_color="FFC300", end_color="FFC300", fill_type="solid")
        elif 0 <= temperature < 15:
            temp_cell.fill = PatternFill(start_color="DAF7A6", end_color="DAF7A6", fill_type="solid")
        elif temperature < 0:
            temp_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")

    wb.save(FILE_PATH)
    wb.close()

def read_last_24_rows():
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    return data[-24:] if len(data) >= 24 else []