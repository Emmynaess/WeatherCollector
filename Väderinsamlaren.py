import requests
import os
import datetime
from datetime import datetime, timedelta, date
import json
import openpyxl
from openpyxl import Workbook, load_workbook
import time


# Koordinater
lat = 59.30
lon = 18.02

# API
URL = f'https://opendata-download-metfcst.smhi.se/api/category/pmp3g/version/2/geotype/point/lon/{lon}/lat/{lat}/data.json'

def get_data_from_smhi():
    return requests.get(URL)

# DATUM
def get_dates(smhi):
    smhi = smhi.json()
    viktig_data = smhi['timeSeries']
    date_data = []

    for temp in viktig_data[:24]:
        timeText = temp['validTime']
        timeText = timeText.replace('T', ' ')
        timeText = timeText[:-1]
        date = datetime.strptime(timeText, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        date_data.append(date)

    return date_data
date_smhi = get_dates(get_data_from_smhi())

# TID
def get_timestamp(smhi):
    smhi = smhi.json()
    viktig_data = smhi['timeSeries']
    time_data = []

    for temp in viktig_data[:24]:
        timeText = temp['validTime']
        timeText = timeText.replace('T', ' ')
        timeText = timeText[:-1]
        time = datetime.strptime(timeText, '%Y-%m-%d %H:%M:%S')
        time += timedelta(hours=1)
        time_data.append(time.strftime('%H:%M:%S'))

    return time_data
time_smhi = get_timestamp(get_data_from_smhi())

# TEMPERATURES
def get_temperatures(smhi):
    smhi = smhi.json()
    viktig_data = smhi['timeSeries']
    temp_data = []

    for temp in viktig_data[:24]:
        for parameter in temp['parameters']:
            if parameter['unit'] == 'Cel':
                t = parameter['values'][0]
                temp_data.append(float(t))
    return temp_data
temp_smhi = get_temperatures(get_data_from_smhi())

# NEDERBÖRD
def get_RainOrSnow(smhi) -> bool:
    smhi = smhi.json()
    viktig_data = smhi['timeSeries']
    rainorsnow_data = []

    for data in viktig_data[:24]:
        for pc in data['parameters']:
            if pc['name'] == 'pcat':
                if pc['values'][0] == 0:
                    rainorsnow_data.append('False')
                else:
                    rainorsnow_data.append('True')
    return rainorsnow_data
rainorsnow_smhi = get_RainOrSnow(get_data_from_smhi())

def create_excel_file_if_not_exists():
    file_path = 'Väderdata.xlsx'
    if not os.path.isfile(file_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Created'
        ws['B1'] = 'Longitude'
        ws['C1'] = 'Latitude'
        ws['D1'] = 'Datum'
        ws['E1'] = 'Hour'
        ws['F1'] = 'Temperature'
        ws['G1'] = 'Rain or snow'
        ws['H1'] = 'Provider'
        wb.save(file_path)
        wb.close()

def run():
   
    while True:
        print("MENY")
        print("1. Hämta senaste data")
        print("2. Skriv ut prognos")
        print("9. Avsluta")
   
        val = input("Välj ett alternativ: ")
   
        if val == '1':
            created = datetime.now()
            data_from_smhi = get_data_from_smhi()
            dates_smhi = get_dates(data_from_smhi)
            time_smhi = get_timestamp(data_from_smhi)
            temp_smhi = get_temperatures(data_from_smhi)
            rainorsnow_smhi = get_RainOrSnow(data_from_smhi)
            create_excel_file_if_not_exists()
            print('\nData har hämtats.\n')

            try:
                wb = load_workbook('Väderdata.xlsx')
                ws = wb.active
                for i in range(24):
                    ws.append([created, lon, lat, dates_smhi[i], time_smhi[i], temp_smhi[i], rainorsnow_smhi[i], 'SMHI'])
                wb.save('Väderdata.xlsx')
                wb.close()
                time.sleep(1)
                print('\nPrognosdata har skrivits till Excel-filen.\n')
            except FileNotFoundError:
                print('\nFilen Väderdata.xlsx hittades inte eller kunde inte öppnas.\n')
               

        elif val == '2':
            time.sleep(1)
            nu = datetime.now()
            file_path = 'Väderdata.xlsx'
           
            try:
                wb = load_workbook(file_path)
                sheet = wb.active
                totala_rows = sheet.max_row

                temperatures = []
                hours = []
                rain_or_snow = []

                for row in range(totala_rows - 23, totala_rows + 1):
                    temperature = None
                    hour = None
                    rain_snow = None

                    for column in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=row, max_row=row):
                        for cell in column:
                            if column[0].column == 6:
                                temperature = cell.value
                            elif column[0].column == 5:
                                hour = cell.value
                            elif column[0].column == 7:
                                rain_snow = cell.value

                    temperatures.append(temperature)
                    hours.append(hour)
                    rain_or_snow.append(rain_snow)

                if len(temperatures) >= 24 and len(hours) >= 24 and len(rain_or_snow) >= 24:
                    print(f'\nPrognos från SMHI {nu.strftime("%Y-%m-%d")}:')
                    print("\x1b[34m{:<15} {:<15} {:<15}\x1b[0m".format("Temperature", "Hour", "Rain/Snow"))
                    for i in range(24):
                        rain_snow_nederbörd = 'Nederbörd' if rain_or_snow[i] == 'True' else 'Ingen nederbörd'
                        print("{:<15} {:<15} {:<15}".format(temperatures[i], hours[i], rain_snow_nederbörd))
                else:
                    print('Det finns inte tillräckligt med data i Excel-filen (färre än 24 rader).')
                wb.close()
            except FileNotFoundError:
                print(f'\nFilen {file_path} hittades inte.\n')
               
        elif val == '9':
            print('Avslutar programmet.')
            break
        else:
            print('Ogiltligt val. Välj igen')

if __name__ == '__main__':
    run()
